import csv
from schedule import every, repeat, run_pending
from time import sleep
from bs4 import BeautifulSoup
import requests
from datetime import datetime
from openpyxl import load_workbook
from sqlalchemy import (
    create_engine,
    MetaData,
    Table,
    Column,
    Integer,
    String,
    Float,
)

engine = create_engine("sqlite:///logs.db", echo=True)
meta = MetaData()

logs = Table(
    "logs",
    meta,
    Column("id", Integer, primary_key=True),
    Column("branch", String(80)),
    Column("brand", String(80)),
    Column("date", String(80)),
    Column("hour", String(80)),
    Column("status", String(80)),
    Column("current_rating", Float),
)
meta.create_all(engine)


def save_to_db(branch, brand, date, hour, status, current_rating):

    sql = logs.insert().values(
        branch=branch,
        brand=brand,
        date=date,
        hour=hour,
        status=status,
        current_rating=current_rating,
    )
    engine.execute(sql)


def save_to_db_log(branch, brand, date, hour, log):

    sql = logs.insert().values(
        branch=branch, brand=brand, date=date, hour=hour, status=log, current_rating=log
    )


def save_to_excel(
    ws, wb, myFileName, branch, brand, date, hour, status, current_rating
):
    ws.insert_rows(1)
    ws.cell(row=1, column=1, value=branch)
    ws.cell(row=1, column=2, value=brand)
    ws.cell(row=1, column=3, value=date)
    ws.cell(row=1, column=4, value=hour)
    ws.cell(row=1, column=5, value=status)
    ws.cell(row=1, column=6, value=current_rating)

    # ws.append([branch, brand, date, hour, status, current_rating])

    wb.save(filename=myFileName)

    wb.close()


def save_to_excel_log(ws, wb, myFileName, branch, brand, date, hour, log):


    ws.insert_rows(1)
    ws.cell(row=1, column=1, value=branch)
    ws.cell(row=1, column=2, value=brand)
    ws.cell(row=1, column=3, value=date)
    ws.cell(row=1, column=4, value=hour)
    ws.cell(row=1, column=5, value=log)
    ws.cell(row=1, column=6, value=log)

    # ws.append([branch, brand, date, hour, log, log])

    wb.save(filename=myFileName)

    wb.close()

def save_to_csv(branch,brand,date,hour,status,current_rating):
    None
def save_to_csv_log():
    None
def getir_to_excel_first():

    myFileName = r"./logs.xlsx"

    wb = load_workbook(filename=myFileName)

    # TODO: gonna create the page for every restaurant and make the changes into those restaurants
    ws = wb["Sheet1"]

    # TODO: Be sure that all the restaurant are open! And get the links. Then create a if statement to make OPEN/CLOSE diff. Belove statement is not quite right, gotta change it with the real time case.
    url_dict = {
        "Ac??badem Alle Bowls": "https://getir.com/yemek/restoran/alle-bowls-acibadem-mah-kadikoy-istanbul/",
        "Ac??badem Arianas Cheesecake": "https://getir.com/yemek/restoran/ariana-s-cheesecake-acibadem-mah-kadikoy-istanbul/",
        "Ac??badem Big Bold Quick": "https://getir.com/yemek/restoran/bbq-big-bold-quick-acibadem-mah-kadikoy-istanbul",
        "Ac??badem Caesar Salad By": "https://getir.com/yemek/restoran/caesar-salad-by-chef-amadeo-acibadem-mah-kadikoy-istanbul",
        "Ac??badem ??osa": "https://getir.com/yemek/restoran/cosa-bi-corba-bi-salata-acibadem-mah-kadikoy-istanbul/",
        "Ac??badem Detroit Bad Boys Pizza": "https://getir.com/yemek/restoran/detroit-bad-boys-pizza-acibadem-mah-kadikoy-istanbul",
        "Ac??badem Dlycious Dyssert": "https://getir.com/yemek/restoran/dydy-dylicious-dyssert-acibadem-mah-kadikoy-istanbul",
        "Ac??badem Doyuyo": "https://getir.com/yemek/restoran/doyuyo-sarayardi-cad-kadikoy-istanbul/",
        "Ac??badem El Pollo Lasso": "https://getir.com/yemek/restoran/el-pollo-lasso-acibadem-mah-kadikoy-istanbul/",
        "Ac??badem Eti??ler K??fte": "https://getir.com/yemek/restoran/et-isleri-kofte-burger-durum-acibadem-mah-kadikoy-istanbul/",
        "Ac??badem Fadelini": "https://getir.com/yemek/restoran/fadelini-acibadem-mah-kadikoy-istanbul/",
        "Ac??badem Fun For Fit": "https://getir.com/yemek/restoran/fun-for-fit-acibadem-mah-kadikoy-istanbul",
        "Ac??badem G&G Burger": "https://getir.com/yemek/restoran/g-g-burger-acibadem-mah-kadikoy-istanbul-2/",
        "Ac??badem Gurra Tavuk": "https://getir.com/yemek/restoran/gurra-tavuk-doner-acibadem-mah-kadikoy-istanbul",
        "Ac??badem Jay Jay Fries": "https://getir.com/yemek/restoran/jay-jay-fries-acibadem-mah-kadikoy-istanbul",
        "Ac??badem Kale Arkas?? Mutfak": "https://getir.com/yemek/restoran/kale-arkasi-mutfak-acibadem-mah-kadikoy-istanbul",
        "Ac??badem Kengeres ??i?? K??fte": "https://getir.com/yemek/restoran/kengeres-gurme-cig-kofte-acibadem-mah-kadikoy-istanbul",
        "Ac??badem Madritas": "https://getir.com/yemek/restoran/madritas-acibadem-mah-kadikoy-istanbul",
        "Ac??badem Mztps Meze": "https://getir.com/yemek/restoran/mztps-meze-tapas-acibadem-mah-kadikoy-istanbul",
        "Ac??badem Nane Mant??": "https://getir.com/yemek/restoran/nane-manti-acibadem-mah-kadikoy-istanbul/",
        "Ac??badem Noody": "https://getir.com/yemek/restoran/noody-acibadem-mah-kadikoy-istanbul/",
        "Ac??badem Red Haag": "https://getir.com/yemek/restoran/red-haag-cafe-brasserie-acibadem-mah-kadikoy-istanbul",
        "Ac??badem Rylee's Ranch Salad": "https://getir.com/yemek/restoran/rylee-s-ranch-salad-acibadem-mah-kadikoy-istanbul/",
        "Ac??badem Seez Beez": "https://getir.com/yemek/restoran/seez-beez-falafel-wraps-burgers-acibadem-mah-kadikoy-istanbul",
        "Ac??badem Senor Torreon": "https://getir.com/yemek/restoran/senor-torreon-red-hot-spicy-food-acibadem-mah-kadikoy-istanbul/",
        "Ac??badem Tabur K??fte": "https://getir.com/yemek/restoran/tabur-kofte-acibadem-mah-kadikoy-istanbul",
        "Ac??badem The Bowl": "https://getir.com/yemek/restoran/the-bowl-best-of-we-love-acibadem-mah-kadikoy-istanbul",
        "Ac??badem Veganista": "https://getir.com/yemek/restoran/veganista-acibadem-mah-kadikoy-istanbul",
        "Ata??ehir Alle Bowls": "https://getir.com/yemek/restoran/alle-bowls-ataturk-mah-atasehir-istanbul/",
        "Ata??ehir Arianas Cheesecake": "https://getir.com/yemek/restoran/ariana-s-cheesecake-ataturk-mah-atasehir-istanbul/",
        "Ata??ehir Big Bold Quick": "https://getir.com/yemek/restoran/bbq-big-bold-quick-ataturk-mah-atasehir-istanbul/",
        "Ata??ehir Caesar Salad By": "https://getir.com/yemek/restoran/caesar-salad-by-chef-amadeo-ataturk-mah-atasehir-istanbul",
        "Ata??ehir ??osa": "https://getir.com/yemek/restoran/cosa-bi-corba-bi-salata-ataturk-mah-atasehir-istanbul/",
        "Ata??ehir Detroit Bad Boys Pizza": "https://getir.com/yemek/restoran/detroit-bad-boys-pizza-ataturk-mah-atasehir-istanbul",
        "Ata??ehir Dlycious Dyssert": "https://getir.com/yemek/restoran/dydy-dylicious-dyssert-ataturk-mah-atasehir-istanbul",
        "Ata??ehir Doyuyo": "https://getir.com/yemek/restoran/doyuyo-ataturk-mah-atasehir-istanbul/",
        "Ata??ehir El Pollo Lasso": "https://getir.com/yemek/restoran/el-pollo-lasso-ataturk-mah-atasehir-istanbul/",
        "Ata??ehir Fadelini": "https://getir.com/yemek/restoran/fadelini-ataturk-mah-atasehir-istanbul/",
        "Ata??ehir Fun For Fit": " https://getir.com/yemek/restoran/fun-for-fit-ataturk-mah-atasehir-istanbul",
        "Ata??ehir Gurra Tavuk": "https://getir.com/yemek/restoran/gurra-tavuk-doner-ataturk-mah-atasehir-istanbul",
        "Ata??ehir Jay Jay Fries": "https://getir.com/yemek/restoran/jay-jay-fries-ataturk-mah-atasehir-istanbul",
        "Ata??ehir Kale Arkas?? Mutfak": " https://getir.com/yemek/restoran/kale-arkasi-mutfak-ataturk-mah-atasehir-istanbul",
        "Ata??ehir Kengeres ??i?? K??fte": "https://getir.com/yemek/restoran/kengeres-gurme-cig-kofte-ataturk-mah-atasehir-istanbul",
        "Ata??ehir Madritas": "https://getir.com/yemek/restoran/madritas-ataturk-mah-atasehir-istanbul/",
        "Ata??ehir Mztps Meze": " https://getir.com/yemek/restoran/mztps-meze-tapas-ataturk-mah-atasehir-istanbul",
        "Ata??ehir Nane Mant??": " https://getir.com/yemek/restoran/nane-manti-ataturk-mah-atasehir-istanbul/",
        "Ata??ehir Noody": "https://getir.com/yemek/restoran/noody-ataturk-mah-atasehir-istanbul/",
        "Ata??ehir Red Haag": "https://getir.com/yemek/restoran/red-haag-cafe-brasserie-ataturk-mah-atasehir-istanbul/",
        "Ata??ehir Rylee's Ranch Salad": "https://getir.com/yemek/restoran/rylee-s-ranch-salad-ataturk-mah-atasehir-istanbul/",
        "Ata??ehir Seez Beez": "https://getir.com/yemek/restoran/seez-beez-falafel-wraps-burgers-ataturk-mah-atasehir-istanbul",
        "Ata??ehir Senor Torreon": "https://getir.com/yemek/restoran/senor-torreon-red-hot-spicy-food-ataturk-mah-atasehir-istanbul/",
        "Ata??ehir Sushi Master": "https://getir.com/yemek/restoran/sushi-master-ataturk-mah-atasehir-istanbul/",
        "Ata??ehir The Bowl": "https://getir.com/yemek/restoran/the-bowl-best-of-we-love-ataturk-mah-atasehir-istanbul",
        "Ata??ehir Veganista": "https://getir.com/yemek/restoran/veganista-ataturk-mah-atasehir-istanbul",
        "Kozyata???? Ali Veli Gurme Pide": "https://getir.com/yemek/restoran/ali-veli-gurme-pide-kozyatagi-mah-kadikoy-istanbul/",
        "Kozyata???? Alle Bowls": "https://getir.com/yemek/restoran/alle-bowls-kozyatagi-mah-kadikoy-istanbul/",
        "Kozyata???? Arianas Cheesecake": "https://getir.com/yemek/restoran/ariana-s-cheesecake-kozyatagi-mah-kadikoy-istanbul/",
        "Kozyata???? Big Bold Quick": "https://getir.com/yemek/restoran/bbq-big-bold-quick-kozyatagi-mah-kadikoy-istanbul",
        "Kozyata???? Caesar Salad By": "https://getir.com/yemek/restoran/caesar-salad-by-chef-amadeo-kozyatagi-mah-kadikoy-istanbul",
        "Kozyata???? ??osa": "https://getir.com/yemek/restoran/cosa-bi-corba-bi-salata-kozyatagi-mah-kadikoy-istanbul/",
        "Kozyata???? Detroit Bad Boys Pizza": "https://getir.com/yemek/restoran/detroit-bad-boys-pizza-kozyatagi-mah-kadikoy-istanbul",
        "Kozyata???? Dlycious Dyssert": "https://getir.com/yemek/restoran/dydy-dylicious-dyssert-kozyatagi-mah-kadikoy-istanbul",
        "Kozyata???? Doyuyo": "https://getir.com/yemek/restoran/doyuyo-kozyatagi-mah-kadikoy-istanbul/",
        "Kozyata???? El Pollo Lasso": "https://getir.com/yemek/restoran/el-pollo-lasso-kozyatagi-mah-kadikoy-istanbul//",
        "Kozyata???? Eti??ler K??fte": "https://getir.com/yemek/restoran/et-isleri-kofte-burger-durum-kozyatagi-mah-kadikoy-istanbul/",
        "Kozyata???? Fadelini": "https://getir.com/yemek/restoran/fadelini-kozyatagi-mah-kadikoy-istanbul/",
        "Kozyata???? Fun For Fit": "https://getir.com/yemek/restoran/fun-for-fit-kozyatagi-mah-kadikoy-istanbul",
        "Kozyata???? G&G Burger": "https://getir.com/yemek/restoran/g-g-burger-kozyatagi-mah-kadikoy-istanbul/",
        "Kozyata???? Gurra Tavuk": "https://getir.com/yemek/restoran/gurra-tavuk-doner-kozyatagi-mah-kadikoy-istanbul",
        "Kozyata???? Jay Jay Fries": "https://getir.com/yemek/restoran/jay-jay-fries-kozyatagi-mah-kadikoy-istanbul",
        "Kozyata???? Kale Arkas?? Mutfak": "https://getir.com/yemek/restoran/kale-arkasi-mutfak-kozyatagi-mah-kadikoy-istanbul",
        "Kozyata???? Kengeres ??i?? K??fte": "https://getir.com/yemek/restoran/kengeres-gurme-cig-kofte-kozyatagi-mah-kadikoy-istanbul",
        "Kozyata???? Madritas": "https://getir.com/yemek/restoran/madritas-kozyatagi-mah-kadikoy-istanbul",
        "Kozyata???? Mztps Meze": "https://getir.com/yemek/restoran/mztps-meze-tapas-kozyatagi-mah-kadikoy-istanbul",
        "Kozyata???? Nane Mant??": "https://getir.com/yemek/restoran/nane-manti-kozyatagi-mah-kadikoy-istanbul/",
        "Kozyata???? Noody": "https://getir.com/yemek/restoran/noody-kozyatagi-mah-kadikoy-istanbul/",
        "Kozyata???? Red Haag": "https://getir.com/yemek/restoran/red-haag-cafe-brasserie-kozyatagi-mah-kadikoy-istanbul",
        "Kozyata???? Rylee's Ranch Salad": "https://getir.com/yemek/restoran/rylee-s-ranch-salad-kozyatagi-mah-kadikoy-istanbul/",
        "Kozyata???? Senor Torreon": "https://getir.com/yemek/restoran/senor-torreon-red-hot-spicy-food-kozyatagi-mah-kadikoy-istanbul/",
        "Kozyata???? Sushi Master": "https://getir.com/yemek/restoran/sushi-master-kozyatagi-mah-kadikoy-istanbul/",
        "Kozyata???? Tabur K??fte": "https://getir.com/yemek/restoran/tabur-kofte-kozyatagi-mah-kadikoy-istanbul",
        "Kozyata???? The Bowl": "https://getir.com/yemek/restoran/the-bowl-best-of-we-love-kozyatagi-mah-kadikoy-istanbul",
        "Kozyata???? Veganista": "https://getir.com/yemek/restoran/veganista-kozyatagi-mah-kadikoy-istanbul",
        "FSM Alle Bowls": "https://getir.com/yemek/restoran/alle-bowls-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "FSM Arianas Cheesecake": "https://getir.com/yemek/restoran/ariana-s-cheesecake-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "FSM Big Bold Quick": "https://getir.com/yemek/restoran/bbq-big-bold-quick-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "FSM Caesar Salad By": "https://getir.com/yemek/restoran/caesar-salad-by-chef-amadeo-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "FSM ??osa": "https://getir.com/yemek/restoran/cosa-bi-corba-bi-salata-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "FSM Detroit Bad Boys Pizza": "https://getir.com/yemek/restoran/detroit-bad-boys-pizza-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "FSM Dlycious Dyssert": "https://getir.com/yemek/restoran/dydy-dylicious-dyssert-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "FSM El Pollo Lasso": "https://getir.com/yemek/restoran/el-pollo-lasso-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "FSM Eti??ler K??fte": "https://getir.com/yemek/restoran/et-isleri-kofte-burger-durum-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "FSM Fadelini": "https://getir.com/yemek/restoran/fadelini-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "FSM Fun For Fit": "https://getir.com/yemek/restoran/fun-for-fit-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "FSM G&G Burger": "https://getir.com/yemek/restoran/g-g-burger-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "FSM Gurra Tavuk": "https://getir.com/yemek/restoran/gurra-tavuk-doner-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "FSM Kale Arkas?? Mutfak": "https://getir.com/yemek/restoran/kale-arkasi-mutfak-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "FSM Kengeres ??i?? K??fte": "https://getir.com/yemek/restoran/kengeres-gurme-cig-kofte-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "FSM Madritas": "https://getir.com/yemek/restoran/madritas-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "FSM Mztps Meze": "https://getir.com/yemek/restoran/mztps-meze-tapas-kozyatagi-mah-kadikoy-istanbul",
        "FSM Nane Mant??": "https://getir.com/yemek/restoran/nane-manti-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "FSM Noody": "https://getir.com/yemek/restoran/noody-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "FSM Red Haag": "https://getir.com/yemek/restoran/red-haag-cafe-brasserie-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "FSM Rylee's Ranch Salad": "https://getir.com/yemek/restoran/rylee-s-ranch-salad-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "FSM Senor Torreon": "https://getir.com/yemek/restoran/senor-torreon-red-hot-spicy-food-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "FSM Sushi Master": "https://getir.com/yemek/restoran/sushi-master-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "FSM Tabur K??fte": "https://getir.com/yemek/restoran/tabur-kofte-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "FSM The Bowl": "https://getir.com/yemek/restoran/the-bowl-best-of-we-love-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "FSM Veganista": "https://getir.com/yemek/restoran/veganista-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "Maslak Alle Bowls": "https://getir.com/yemek/restoran/alle-bowls-maslak-mah-sariyer-istanbul/",
        "Maslak Arianas Cheesecake": "https://getir.com/yemek/restoran/ariana-s-cheesecake-maslak-mah-sariyer-istanbul/",
        "Maslak Big Bold Quick": "https://getir.com/yemek/restoran/bbq-big-bold-quick-maslak-mah-sariyer-istanbul/",
        "Maslak Caesar Salad By": "https://getir.com/yemek/restoran/caesar-salad-by-chef-amadeo-maslak-mah-sariyer-istanbul",
        "Maslak ??osa": "https://getir.com/yemek/restoran/cosa-bi-corba-bi-salata-maslak-mah-sariyer-istanbul/",
        "Maslak Detroit Bad Boys Pizza": "https://getir.com/yemek/restoran/detroit-bad-boys-pizza-maslak-mah-sariyer-istanbul",
        "Maslak Dlycious Dyssert": "https://getir.com/yemek/restoran/dydy-dylicious-dyssert-maslak-mah-sariyer-istanbul",
        "Maslak El Pollo Lasso": "https://getir.com/yemek/restoran/el-pollo-lasso-maslak-mah-sariyer-istanbul/",
        "Maslak Eti??ler K??fte": "https://getir.com/yemek/restoran/et-isleri-kofte-burger-durum-maslak-mah-sariyer-istanbul/",
        "Maslak Fadelini": "https://getir.com/yemek/restoran/fadelini-maslak-mah-sariyer-istanbul/",
        "Maslak Fun For Fit": "https://getir.com/yemek/restoran/fun-for-fit-maslak-mah-sariyer-istanbul",
        "Maslak G&G Burger": "https://getir.com/yemek/restoran/g-g-burger-maslak-mah-sariyer-istanbul/",
        "Maslak Gurra Tavuk": "https://getir.com/yemek/restoran/gurra-tavuk-doner-maslak-mah-sariyer-istanbul",
        "Maslak Jay Jay Fries": "https://getir.com/yemek/restoran/jay-jay-fries-maslak-mah-sariyer-istanbul",
        "Maslak Kale Arkas?? Mutfak": "https://getir.com/yemek/restoran/kale-arkasi-mutfak-maslak-mah-sariyer-istanbul",
        "Maslak Kengeres ??i?? K??fte": "https://getir.com/yemek/restoran/kengeres-gurme-cig-kofte-maslak-mah-sariyer-istanbul",
        "Maslak Madritas": "https://getir.com/yemek/restoran/madritas-maslak-mah-sariyer-istanbul/",
        "Maslak Mztps Meze": "https://getir.com/yemek/restoran/mztps-meze-tapas-maslak-mah-sariyer-istanbul",
        "Maslak Nane Mant??": "https://getir.com/yemek/restoran/nane-manti-evi-maslak-mah-sariyer-istanbul/",
        "Maslak Noody": "https://getir.com/yemek/restoran/noody-maslak-mah-sariyer-istanbul/",
        "Maslak Red Haag": "https://getir.com/yemek/restoran/red-haag-cafe-brasserie-maslak-mah-sariyer-istanbul/",
        "Maslak Rylee's Ranch Salad": "https://getir.com/yemek/restoran/rylee-s-ranch-salad-maslak-mah-sariyer-istanbul/",
        "Maslak Senor Torreon": "https://getir.com/yemek/restoran/senor-torreon-red-hot-spicy-food-maslak-mah-sariyer-istanbul/",
        "Maslak Sushi Master": "https://getir.com/yemek/restoran/sushi-master-maslak-mah-sariyer-istanbul/",
        "Maslak Tabur K??fte": "https://getir.com/yemek/restoran/tabur-kofte-maslak-mah-sariyer-istanbul/",
        "Maslak The Bowl": "https://getir.com/yemek/restoran/the-bowl-best-of-we-love-maslak-mah-sariyer-istanbul",
        "Maslak Veganista": "https://getir.com/yemek/restoran/veganista-maslak-mah-sariyer-istanbul",
        "??zmir Alle Bowls": "https://getir.com/yemek/restoran/alle-bowls-cigli-atasehir-mah-cigli-izmir/",
        "??zmir Arianas Cheesecake": "https://getir.com/yemek/restoran/ariana-s-cheesecake-cigli-atasehir-mah-cigli-izmir/",
        "??zmir Big Bold Quick": "https://getir.com/yemek/restoran/bbq-big-bold-quick-atasehir-mah-cigli-izmir/",
        "??zmir Caesar Salad By": "https://getir.com/yemek/restoran/caesar-salad-by-chef-amadeo-cigli-atasehir-mah-cigli-izmir",
        "??zmir ??osa": "https://getir.com/yemek/restoran/cosa-bi-corba-bi-salata-cigli-atasehir-mah-cigli-izmir/",
        "??zmir El Pollo Lasso": "https://getir.com/yemek/restoran/el-pollo-lasso-cigli-atasehir-mah-cigli-izmir/",
        "??zmir Eti??ler K??fte": "https://getir.com/yemek/restoran/etisleri-kofte-burger-durum-cigli-atasehir-mah-cigli-izmir/",
        "??zmir Fadelini": "https://getir.com/yemek/restoran/fadelini-cigli-atasehir-mah-cigli-izmir/",
        "??zmir Fun For Fit": "https://getir.com/yemek/restoran/fun-for-fit-maslak-mah-sariyer-istanbul",
        "??zmir Gurra Tavuk": "https://getir.com/yemek/restoran/gurra-tavuk-doner-cigli-atasehir-mah-cigli-izmir",
        "??zmir Jay Jay Fries": "https://getir.com/yemek/restoran/jay-jay-fries-maslak-mah-sariyer-istanbul",
        "??zmir Kale Arkas?? Mutfak": "https://getir.com/yemek/restoran/kale-arkasi-mutfak-cigli-atasehir-mah-cigli-izmir",
        "??zmir Kengeres ??i?? K??fte": "https://getir.com/yemek/restoran/kengeres-gurme-cig-kofte-cigli-atasehir-mah-cigli-izmir",
        "??zmir Madritas": "https://getir.com/yemek/restoran/madritas-cigli-atasehir-mah-cigli-izmir/",
        "??zmir Mztps Meze": "https://getir.com/yemek/restoran/mztps-meze-tapas-cigli-atasehir-mah-cigli-izmir",
        "??zmir Nane Mant??": "https://getir.com/yemek/restoran/nane-manti-cigli-atasehir-mah-cigli-izmir/",
        "??zmir Noody": "https://getir.com/yemek/restoran/noody-cigli-atasehir-mah-cigli-izmir/",
        "??zmir Red Haag": "https://getir.com/yemek/restoran/red-haag-cafe-brasserie-atasehir-mah-cigli-izmir/",
        "??zmir Rylee's Ranch Salad": "https://getir.com/yemek/restoran/rylee-s-ranch-salad-cigli-atasehir-mah-cigli-izmir/",
        "??zmir Seez Beez": "https://getir.com/yemek/restoran/seez-beez-falafel-burgers-wraps-cigli-atasehir-mah-cigli-izmir/",
        "??zmir Senor Torreon": "https://getir.com/yemek/restoran/senor-torreon-red-hot-spicy-food-cigli-atasehir-mah-cigli-izmir/",
        "??zmir Tabur K??fte": "https://getir.com/yemek/restoran/tabur-kofte-cigli-atasehir-mah-cigli-izmir/",
        "??zmir The Bowl": "https://getir.com/yemek/restoran/the-bowl-best-of-we-love-cigli-atasehir-mah-cigli-izmir",
        "??zmir Veganista": "https://getir.com/yemek/restoran/veganista-cigli-atasehir-mah-cigli-izmir"
    }

    num = str(len(list(url_dict.keys())))
    print(f"{num} restaurant's data are gonna be collected.")

    for url_key in url_dict.keys():
        url = url_dict[url_key]
        url_key_list = url_key.split(" ")
        branch = url_key_list[0]
        url_key_list.pop(0)
        brand = " ".join(url_key_list)
        now = datetime.now()
        format_date = "%d/%m/%Y"
        format_hour = "%H:%M:%S"
        # format datetime using strftime()
        date = now.strftime(format_date)
        hour = now.strftime(format_hour)
        r = None

        index = str(list(url_dict.keys()).index(url_key))

        # This is for 'if connection get suspened by getir.com because we are making so much request in a short time'
        try:
            r = requests.get(url)
        except Exception as e:
            log = str(e)

            save_to_excel_log(
                ws=ws,
                wb=wb,
                myFileName=myFileName,
                branch=branch,
                brand=brand,
                date=date,
                hour=hour,
                log=log,
            )

            save_to_db_log(branch=branch, brand=brand, date=date,hour=hour, log=log)

            print(index + log)

            continue

        if r == None:
            print(index + "URL is None")
            continue

        # TODO: This is gonna change, because we are not checking the restauant open/close status by status.code normally. # Have the screenshot in the screenshots.

        soup = BeautifulSoup(r.content, "html.parser")

        try:
            close = get_close(soup=soup)

        except Exception as e:
            log = str(e)

            save_to_excel_log(
                ws=ws,
                wb=wb,
                myFileName=myFileName,
                branch=branch,
                brand=brand,
                date=date,
                hour=hour,
                log=log,
            )

            save_to_db_log(branch=branch, brand=brand, date=date, hour=hour, log=log)

            print(index + log)
            continue

        try:
            current_rating = get_rating(soup=soup)

        except Exception as e:
            current_rating = None

        # Check that if close state is exsist in the html.
        if close == None:
            # TODO: append the date to the current sheet. But this is gonna change by the restaurant.
            status = "A??IK"

            try:

                save_to_excel(
                    ws=ws,
                    wb=wb,
                    myFileName=myFileName,
                    branch=branch,
                    brand=brand,
                    date=date,
                    hour=hour,
                    status=status,
                    current_rating=current_rating,
                )

                print(index + "Changes saved to excel!")

            except Exception as e:
                print(index + str(e))
                print(index + "Couldn't write to excel! Keep looping")
                continue

            try:
                save_to_db(
                    branch=branch,
                    brand=brand,
                    date=date,
                    hour=hour,
                    status=status,
                    current_rating=current_rating,
                )

            except Exception as e:
                print(index + str(e))
                print("Couldn't write to db!")
                continue

        # TODO: This is gonna be canceled because we are not going to calculate the current OPEN/CLOSE status with this
        else:
            status = "KAPALI"

            try:

                save_to_excel(
                    ws=ws,
                    wb=wb,
                    myFileName=myFileName,
                    branch=branch,
                    brand=brand,
                    date=date,
                    hour=hour,
                    status=status,
                    current_rating=current_rating,
                )

                print(index + "Changes saved to excel!")

            except Exception as e:

                print(index + str(e))
                print(index + "Couldn't write to excel! Keep looping")
                continue

            try:
                save_to_db(
                    branch=branch,
                    brand=brand,
                    date=date,
                    hour=hour,
                    status=status,
                    current_rating=current_rating,
                )

            except Exception as e:
                print(index + str(e))
                print("Couldn't write to db!")
                continue

        sleep(5)

    return url_dict  # temp


def get_rating(soup):
    # Find the rate
    body = soup.body
    s = body.find("div", id="__next")
    s2 = s.find("div", class_="sc-212542e0-2 ckZpLq")
    s4 = s2.find("main", class_="sc-212542e0-0 iiapCb")
    s5 = s4.find("div", class_="sc-e85e5299-0 sc-4e0754cc-0 hkaVQN klYfLJ")
    s6 = s5.find("div", class_="sc-4e0754cc-1 YPsgm")
    s7 = s6.find("div", class_="sc-4e0754cc-3 bwTzrw")
    s8 = s7.find("div", class_="sc-7047f3e2-6 iFptQI")
    s9 = s8.find("div", class_="style__Wrapper-sc-__sc-sbxwka-15 jPuQcd")
    s10 = s9.find("div", class_="style__CardWrapper-sc-__sc-sbxwka-12 ccsSiU")
    s11 = s10.find("div", class_="style__ContentWrapper-sc-__sc-sbxwka-7 emAjmS")
    s12 = s11.find("div", class_="sc-7047f3e2-0 iJHJBI")
    s13 = s12.find("div", class_="sc-7047f3e2-3 hbiBbV")
    rate = s13.find(
        "span", class_="style__Text-sc-__sc-1nwjacj-0 jbOUDC sc-7047f3e2-8 iFDpNz"
    )

    current_rating = rate.get_text()

    return current_rating


def get_close(soup):
    body = soup.body
    s11 = body.find("div", id="__next")
    s10 = s11.find("div", class_="sc-212542e0-2 ckZpLq")
    s9 = s10.find("main", class_="sc-212542e0-0 iiapCb")
    s8 = s9.find("div", class_="sc-e85e5299-0 sc-4e0754cc-0 hkaVQN klYfLJ")
    s7 = s8.find("div", class_="sc-4e0754cc-1 YPsgm")
    s6 = s7.find("div", class_="sc-4e0754cc-3 bwTzrw")
    s5 = s6.find("div", class_="sc-7047f3e2-6 iFptQI")
    s4 = s5.find("div", class_="style__Wrapper-sc-__sc-sbxwka-15 jPuQcd")
    s3 = s4.find("div", class_="style__CardWrapper-sc-__sc-sbxwka-12 ccsSiU")
    s2 = s3.find("div", class_="style__ContentWrapper-sc-__sc-sbxwka-7 emAjmS")
    close = s2.find("div", class_="sc-e27f3f42-0 hPdSRl")

    return close


@repeat(every(1).hour)  # .until("18:30")
def getir_to_excel():

    getir_to_excel_first()


# for the first time and then let the schdule does it's job
getir_to_excel_first()

while True:
    run_pending()
    sleep(1)
